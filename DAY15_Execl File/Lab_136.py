# Reading excel using for loop

import openpyxl
file_path = "C:\\Users\\Welcome\\Desktop\\Demo.xlsx"
wb = openpyxl.load_workbook(file_path)

sh1 = wb["Sheet1"]
row = sh1.max_row
column =sh1.max_column

for i in range (1,row+1):
    for j in range(1,column+1):
        print(sh1.cell(i,j).value)


# import openpyxl
#
# file_path = "C:\\Users\\Welcome\\Desktop\\Demo.xlsx"
# wb = openpyxl.load_workbook(file_path)
#
# sh1 = wb["Sheet1"]
#
# for row in row in sh1.iter_rows(min_row=2,max_row=sh1.max_row,min_col=1,max_col=sh1.max_column):
#     row_values = [cell.value for cell in row]
#     print(row_values)

import openpyxl

file_path = "C:\\Users\\Welcome\\Desktop\\Demo.xlsx"
wb = openpyxl.load_workbook(file_path)

sh1 = wb["Sheet1"]

# Get the dimensions of the sheet
max_row = sh1.max_row
max_col = sh1.max_column

# Iterate through rows, skipping the header (starting from the second row)
for row_index in range(2, max_row + 1):
    row_values = [sh1.cell(row=row_index, column=col_index).value for col_index in range(1, max_col + 1)]
    print(row_values)


