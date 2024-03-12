import openpyxl

file_path = "C:\\Users\\Welcome\\Desktop\\Demo.xlsx"

workbok = openpyxl.load_workbook(file_path)

sheet_name = workbok.sheetnames  # it will show sheet name inside excel
print(sheet_name)
# option 1
sh_name = workbok["Sheet1"]
data = workbok["Sheet1"]["A2"].value
print(data)

# option 2
get_value = sh_name.cell(4, 3).value # using cell method
print(get_value)


